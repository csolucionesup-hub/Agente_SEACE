"""FastAPI dashboard for Agente SEACE.

The web layer intentionally consumes the clean dashboard JSON produced by
``seace_seguimiento.py``. This keeps the product scalable: the crawler/tracking
engine can run as a scheduled worker while the web app serves cached, sanitized
state quickly and safely.
"""

from __future__ import annotations

import hmac
import io
import json
import os
import tempfile
import threading
import time
from datetime import datetime
from dataclasses import asdict
from copy import deepcopy
from pathlib import Path
from typing import Any, Callable

from fastapi import FastAPI, HTTPException, Request
from fastapi.responses import FileResponse, JSONResponse, Response, StreamingResponse
from fastapi.staticfiles import StaticFiles

from seace_commercial_scoring import enrich_opportunity
from seace_relevance import score_relevance
from seace_api import Opportunity, SeaceApiClient
from seace_oportunidades import collect_opportunities
from seace_seguimiento import sync_ocids
from seace_tracking import TrackingStore
from seace_documents import analyze_document, build_technical_file_response, download_verified_document, extract_official_documents, verify_document_link
from seace_conosce import fetch_market_intel
from auth_supabase import SupabaseAuth, bearer_token

DEFAULT_DASHBOARD_PATH = Path("reportes/dashboard-seguimiento.json")
DEFAULT_STATIC_DIR = Path("web")
DEFAULT_TRACKING_DB_PATH = Path("data/seace_tracking.sqlite3")
DEFAULT_SETTINGS_PATH = Path("data/client_settings.json")
DEFAULT_SEARCH_CACHE_TTL = float(os.getenv("LICITASCAN_SEARCH_CACHE_TTL", "600"))
DEFAULT_SEARCH_DEADLINE = float(os.getenv("LICITASCAN_SEARCH_DEADLINE", "20"))
DEFAULT_SEARCH_CACHE_PATH = Path("reportes/search-cache.json")

DEFAULT_SETTINGS: dict[str, Any] = {
    "client_name": "Constructora Andina S.A.C.",
    "business_line": "Construcción e infraestructura",
    "keywords": ["puente", "carretera", "expediente técnico", "obra vial"],
    "negative_keywords": ["hoja de muelle", "abrazadera", "repuesto"],
    # Default para el nicho de puentes/paquetes de puentes: descarta obras chicas y
    # deja las de monto sustancial (S/ 3M+), que son las que suelen incluir puentes.
    "min_amount": 3000000,
    "frequency": "diario",
    "channels": ["Telegram", "WhatsApp", "Google Sheets", "Webhook", "Excel/PDF"],
    "custom_variables": [],
    "ignored_ocids": [],
}

SECURITY_HEADERS = {
    "X-Content-Type-Options": "nosniff",
    "X-Frame-Options": "DENY",
    "Referrer-Policy": "strict-origin-when-cross-origin",
    "Content-Security-Policy": (
        "default-src 'self'; "
        # cdn.jsdelivr.net: librería supabase-js (login con Google) cargada por <script>.
        "script-src 'self' https://cdn.jsdelivr.net; "
        "style-src 'self' 'unsafe-inline'; "
        "img-src 'self' data:; "
        # *.supabase.co: el navegador habla con Supabase Auth (verificar sesión, OAuth).
        "connect-src 'self' https://*.supabase.co; "
        "frame-src https://prodapp2.seace.gob.pe; "
        "frame-ancestors 'none'; "
        "base-uri 'self'; "
        "form-action 'self'"
    ),
}


def _strip_raw_fields(value: Any) -> Any:
    """Return value without raw forensic blobs that should not reach frontend."""
    if isinstance(value, dict):
        return {
            key: _strip_raw_fields(item)
            for key, item in value.items()
            if key not in {"raw", "raw_json", "payload_json"}
        }
    if isinstance(value, list):
        return [_strip_raw_fields(item) for item in value]
    return value


def _empty_dashboard() -> dict[str, Any]:
    return {
        "counts_by_stage": {},
        "counts_by_outcome": {},
        "opportunities": [],
        "recent_events": [],
    }


def load_dashboard(path: str | Path = DEFAULT_DASHBOARD_PATH) -> dict[str, Any]:
    dashboard_path = Path(path)
    if not dashboard_path.exists():
        return _empty_dashboard()
    data = json.loads(dashboard_path.read_text(encoding="utf-8"))
    sanitized = _strip_raw_fields(data)
    sanitized.setdefault("counts_by_stage", {})
    sanitized.setdefault("counts_by_outcome", {})
    sanitized.setdefault("opportunities", [])
    sanitized.setdefault("recent_events", [])
    sanitized["opportunities"] = [enrich_opportunity(item) for item in sanitized["opportunities"]]
    sanitized["opportunities"].sort(
        key=lambda item: (item.get("commercial_score", 0), -(item.get("days_to_critical_date") or 999999)),
        reverse=True,
    )
    return sanitized


def _unique_text_list(values: Any) -> list[str]:
    result: list[str] = []
    for value in values or []:
        text = str(value).strip()
        if text and text not in result:
            result.append(text)
    return result


def normalize_settings(payload: dict[str, Any] | None = None) -> dict[str, Any]:
    merged = {**DEFAULT_SETTINGS, **(payload or {})}
    min_amount = merged.get("min_amount", DEFAULT_SETTINGS["min_amount"])
    try:
        min_amount = max(0, float(min_amount))
    except (TypeError, ValueError):
        min_amount = DEFAULT_SETTINGS["min_amount"]
    if float(min_amount).is_integer():
        min_amount = int(min_amount)
    custom_variables = []
    seen_custom: set[tuple[str, str]] = set()
    for item in merged.get("custom_variables") or []:
        if not isinstance(item, dict):
            continue
        name = str(item.get("name") or "").strip()
        value = str(item.get("value") or "").strip()
        if name and (name, value) not in seen_custom:
            custom_variables.append({"name": name, "value": value})
            seen_custom.add((name, value))
    settings = {
        "client_name": str(merged.get("client_name") or DEFAULT_SETTINGS["client_name"]).strip(),
        "business_line": str(merged.get("business_line") or DEFAULT_SETTINGS["business_line"]).strip(),
        "keywords": _unique_text_list(merged.get("keywords")) or list(DEFAULT_SETTINGS["keywords"]),
        "negative_keywords": _unique_text_list(merged.get("negative_keywords")),
        "min_amount": min_amount,
        "frequency": str(merged.get("frequency") or DEFAULT_SETTINGS["frequency"]).strip(),
        "channels": _unique_text_list(merged.get("channels")),
        "custom_variables": custom_variables,
        "ignored_ocids": _unique_text_list(merged.get("ignored_ocids")),
    }
    return settings


def load_settings(path: str | Path = DEFAULT_SETTINGS_PATH) -> dict[str, Any]:
    settings_path = Path(path)
    if not settings_path.exists():
        return normalize_settings()
    try:
        data = json.loads(settings_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        data = {}
    return normalize_settings(data)


_SETTINGS_LOCK = threading.Lock()


def _atomic_write_text(path: Path, text: str) -> None:
    """Write text via a temp file + atomic replace so readers never see a partial file."""
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp_name = tempfile.mkstemp(dir=str(path.parent), prefix=".settings-", suffix=".tmp")
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as handle:
            handle.write(text)
        os.replace(tmp_name, path)
    except BaseException:
        if os.path.exists(tmp_name):
            os.unlink(tmp_name)
        raise


def save_settings(settings: dict[str, Any], path: str | Path = DEFAULT_SETTINGS_PATH) -> dict[str, Any]:
    normalized = normalize_settings(settings)
    settings_path = Path(path)
    with _SETTINGS_LOCK:
        _atomic_write_text(settings_path, json.dumps(normalized, ensure_ascii=False, indent=2))
    return normalized


def update_settings(
    mutator: Callable[[dict[str, Any]], dict[str, Any]],
    path: str | Path = DEFAULT_SETTINGS_PATH,
) -> dict[str, Any]:
    """Lock-protected read-modify-write so concurrent updates don't lose each other."""
    settings_path = Path(path)
    with _SETTINGS_LOCK:
        normalized = normalize_settings(mutator(load_settings(settings_path)))
        _atomic_write_text(settings_path, json.dumps(normalized, ensure_ascii=False, indent=2))
    return normalized


def _build_timeline(opportunity: dict[str, Any], events: list[dict[str, Any]]) -> list[dict[str, Any]]:
    base = [
        {"event_type": "convocatoria", "title": "Convocatoria", "status": "completed"},
        {"event_type": "seguimiento", "title": "Seguimiento", "status": "current"},
        {"event_type": "buena_pro", "title": "Buena pro", "status": "pending"},
        {"event_type": "contrato", "title": "Contrato", "status": "pending"},
    ]

    if opportunity.get("award_date") or opportunity.get("winner_name"):
        base[2].update(
            {
                "status": "completed",
                "date": opportunity.get("award_date"),
                "description": f"Ganador: {opportunity.get('winner_name', '')}".strip(),
            }
        )
    if opportunity.get("contract_date_signed"):
        base[3].update(
            {
                "status": "completed",
                "date": opportunity.get("contract_date_signed"),
                "description": f"Contrato {opportunity.get('contract_id', '')}".strip(),
            }
        )

    for event in sorted(events, key=lambda item: item.get("occurred_at", "")):
        base.append(
            {
                "event_type": event.get("event_type"),
                "title": event.get("title"),
                "date": event.get("occurred_at"),
                "description": event.get("message"),
                "severity": event.get("severity"),
                "status": "event",
            }
        )
    return base


def _find_opportunity(data: dict[str, Any], ocid: str) -> dict[str, Any]:
    opportunity = next((item for item in data["opportunities"] if item.get("ocid") == ocid), None)
    if opportunity is None:
        raise HTTPException(status_code=404, detail="Opportunity not found")
    return opportunity


def _safe_filename(value: str) -> str:
    cleaned = "".join(char if char.isalnum() or char in {"-", "_"} else "-" for char in value).strip("-")
    while "--" in cleaned:
        cleaned = cleaned.replace("--", "-")
    return cleaned[:80] or "expediente"


def _expediente_export(opportunity: dict[str, Any], events: list[dict[str, Any]]) -> dict[str, Any]:
    return {
        "product": "LicitaScan",
        "export_type": "expediente_seace",
        "opportunity": deepcopy(opportunity),
        "events": deepcopy(events),
        "timeline": _build_timeline(opportunity, events),
    }


def _ficha_payload(opportunity: dict[str, Any]) -> dict[str, Any]:
    source_url = str(opportunity.get("official_source_url") or "https://prodapp2.seace.gob.pe/seacebus-uiwd-pub/buscadorPublico/buscadorPublico.xhtml")
    process_code = str(opportunity.get("process_code") or "")
    entity_name = str(opportunity.get("entity_name") or "")
    return {
        "title": "Ficha de Selección SEACE",
        "source_url": source_url,
        "embed_url": source_url,
        "viewer_mode": "embedded_official_portal",
        "process_code": process_code,
        "entity_name": entity_name,
        "ocid": opportunity.get("ocid"),
        "message": "Abre el Buscador Público SEACE/OECE y usa la nomenclatura para entrar a la Ficha de Selección.",
        "steps": [
            "Abrir Buscador Público SEACE.",
            f"Buscar por nomenclatura: {process_code}" if process_code else "Buscar por nomenclatura del procedimiento.",
            f"Validar entidad: {entity_name}" if entity_name else "Validar la entidad convocante.",
            "En la fila del procedimiento, usar el ícono o enlace ‘Ver Ficha de Selección’.",
            "Dentro de la ficha, revisar Documentos y, si aplica, ‘Ver Expediente Técnico de Obra’."
        ],
        "deep_link_available": False,
        "note": "SEACE usa navegación dinámica; si no acepta enlace profundo estable, LicitaScan abre el buscador con instrucciones exactas.",
    }


def _default_ficha_capture_service(opportunity: dict[str, Any], output_dir: Path) -> dict[str, Any]:
    """Open SEACE with Playwright, try to search/open the ficha, and store evidence."""
    try:
        from playwright.sync_api import sync_playwright
    except Exception as exc:  # noqa: BLE001
        raise RuntimeError("Playwright no está disponible para capturar la ficha SEACE") from exc

    output_dir.mkdir(parents=True, exist_ok=True)
    process_code = str(opportunity.get("process_code") or "").strip()
    entity_name = str(opportunity.get("entity_name") or "").strip()
    source_url = str(opportunity.get("official_source_url") or "https://prodapp2.seace.gob.pe/seacebus-uiwd-pub/buscadorPublico/buscadorPublico.xhtml")
    if not process_code:
        raise RuntimeError("La oportunidad no tiene nomenclatura para buscar en SEACE")

    captured_at = datetime.now().isoformat(timespec="seconds")
    safe_code = _safe_filename(process_code)
    filename = f"ficha-{safe_code}-{datetime.now().strftime('%Y%m%d-%H%M%S')}.png"
    target = output_dir / filename
    steps_completed: list[str] = []
    status = "captured_search_page"
    message = "SEACE fue abierto y se capturó la búsqueda; revisa si la ficha final cargó correctamente."

    old_platform = os.environ.get("PLAYWRIGHT_HOST_PLATFORM_OVERRIDE")
    old_ld_library_path = os.environ.get("LD_LIBRARY_PATH")
    os.environ.setdefault("PLAYWRIGHT_HOST_PLATFORM_OVERRIDE", "ubuntu24.04-x64")
    playwright_libs = "/home/hermesagente2026/.local/playwright-libs/lib"
    if Path(playwright_libs).exists() and playwright_libs not in (old_ld_library_path or ""):
        os.environ["LD_LIBRARY_PATH"] = f"{playwright_libs}:{old_ld_library_path}" if old_ld_library_path else playwright_libs
    try:
        with sync_playwright() as playwright:
            browser = playwright.chromium.launch(headless=True)
            page = browser.new_page(viewport={"width": 1440, "height": 1200})
            page.goto(source_url, wait_until="domcontentloaded", timeout=60000)
            steps_completed.append("open_search")
            page.wait_for_timeout(3000)

            # Confirmed SEACE flow: activate the public procedures tab, search by
            # selection number + bridge keyword/year, then open the ficha icon in
            # the exact row whose nomenclature matches the opportunity.
            selection_number = process_code.split("-")[2] if len(process_code.split("-")) > 2 else ""
            try:
                page.locator('a[href$="tbBuscador:tab1"]').click(force=True, timeout=8000)
                steps_completed.append("open_procedures_tab")
                page.wait_for_timeout(2500)
            except Exception:
                pass

            if selection_number:
                try:
                    page.locator('input[id="tbBuscador:idFormBuscarProceso:numeroSeleccion"]').fill(selection_number, timeout=5000)
                    steps_completed.append("fill_selection_number")
                except Exception:
                    pass
            try:
                search_text = "PUENTE" if "PUENTE" in str(opportunity.get("description", "")).upper() else process_code
                page.locator('input[id="tbBuscador:idFormBuscarProceso:descripcionObjeto"]').fill(search_text, timeout=5000)
                steps_completed.append("fill_description")
            except Exception:
                pass

            try:
                page.locator('button[id="tbBuscador:idFormBuscarProceso:btnBuscarSelToken"]').click(timeout=10000)
                steps_completed.append("search_process")
                page.wait_for_timeout(9000)
            except Exception:
                # fallback: capture current official portal if search button changes
                pass

            ficha_clicked = False
            try:
                rows = page.locator('tbody[id$="dtProcesos_data"] tr')
                for index in range(rows.count()):
                    row = rows.nth(index)
                    if process_code in row.inner_text(timeout=3000):
                        steps_completed.append("match_process_row")
                        ficha_link = row.locator('a:has(img[src*="fichaSeleccion"])').first
                        if ficha_link.is_visible(timeout=5000):
                            ficha_link.click(force=True, timeout=10000)
                            ficha_clicked = True
                            steps_completed.append("open_ficha")
                            page.wait_for_timeout(9000)
                        break
            except Exception:
                pass

            if ficha_clicked:
                status = "captured"
                message = "Ficha de Selección encontrada y capturada automáticamente."
            page.screenshot(path=str(target), full_page=True)
            steps_completed.append("capture")
            browser.close()
    finally:
        if old_platform is None:
            os.environ.pop("PLAYWRIGHT_HOST_PLATFORM_OVERRIDE", None)
        else:
            os.environ["PLAYWRIGHT_HOST_PLATFORM_OVERRIDE"] = old_platform
        if old_ld_library_path is None:
            os.environ.pop("LD_LIBRARY_PATH", None)
        else:
            os.environ["LD_LIBRARY_PATH"] = old_ld_library_path

    return {
        "status": status,
        "message": message,
        "process_code": process_code,
        "entity_name": entity_name,
        "source_url": source_url,
        "image_url": f"/assets/evidencias/{filename}",
        "evidence_path": str(target),
        "captured_at": captured_at,
        "steps_completed": steps_completed,
    }


def _documents_from_record_or_dashboard(api_client: Any, opportunity: dict[str, Any]) -> list[dict[str, Any]]:
    ocid = str(opportunity.get("ocid") or "").strip()
    if api_client and ocid:
        try:
            record = api_client.get_record(ocid)
            documents = extract_official_documents(record)
            if documents:
                return documents
        except Exception:
            pass
    return [deepcopy(document) for document in opportunity.get("official_documents") or []]


def _document_with_status(document: dict[str, Any], *, verify: bool, analyze: bool) -> dict[str, Any]:
    url = str(document.get("download_url") or document.get("url") or "")
    enriched = deepcopy(document)
    enriched.setdefault("safe_to_proxy", url.startswith("https://"))
    enriched["preview_url"] = url
    enriched["download_proxy_url"] = f"/api/documents/download?url={url}" if url else ""
    if verify:
        enriched["verification"] = verify_document_link(url)
    else:
        enriched["verification"] = {"status": "not_checked", "ok": None, "message": "Verifica antes de consumir una descarga."}
    if analyze:
        enriched["analysis"] = analyze_document(url, str(enriched.get("format") or ""))
    return enriched


def _find_document_for_url(api_client: Any, dashboard_path: Path, url: str) -> dict[str, Any] | None:
    data = load_dashboard(dashboard_path)
    for opportunity in data["opportunities"]:
        for document in _documents_from_record_or_dashboard(api_client, opportunity):
            if str(document.get("download_url") or document.get("url") or "") == url:
                return document
    return None


def _parse_keywords(value: str) -> list[str]:
    return [item.strip() for item in value.split(",") if item.strip()]


def _search_result_row(opportunity: Any) -> dict[str, Any]:
    row = asdict(opportunity)
    row["record_url"] = opportunity.record_url
    row["stage"] = "convocado"
    row["outcome"] = "activo"
    row["next_critical_date"] = row.get("tender_end_date")
    return enrich_opportunity(row)


class _TTLCache:
    """Tiny thread-safe TTL cache for expensive search results (per app instance)."""

    def __init__(self, ttl_seconds: float):
        self._ttl = max(0.0, ttl_seconds)
        self._store: dict[Any, tuple[float, Any]] = {}
        self._lock = threading.Lock()

    def get(self, key: Any) -> Any | None:
        if self._ttl <= 0:
            return None
        now = time.monotonic()
        with self._lock:
            entry = self._store.get(key)
            if entry is not None and (now - entry[0]) <= self._ttl:
                return entry[1]
            if entry is not None:
                self._store.pop(key, None)
            return None

    def set(self, key: Any, value: Any) -> None:
        if self._ttl <= 0:
            return
        with self._lock:
            self._store[key] = (time.monotonic(), value)


def _disk_cache_key(cache_key: tuple[Any, ...]) -> str:
    keywords, max_pages, paginate_by = cache_key
    return "||".join(keywords) + f"@@{int(max_pages)}@@{int(paginate_by)}"


def _read_disk_search_cache(
    path: str | Path, cache_key: tuple[Any, ...], ttl_seconds: float
) -> tuple[list[Any], bool] | None:
    """Return cached opportunities from disk if the entry exists and is still fresh."""
    cache_path = Path(path)
    if ttl_seconds <= 0 or not cache_path.exists():
        return None
    try:
        data = json.loads(cache_path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return None
    entry = data.get(_disk_cache_key(cache_key)) if isinstance(data, dict) else None
    if not isinstance(entry, dict):
        return None
    timestamp = entry.get("ts")
    if not isinstance(timestamp, (int, float)) or (time.time() - timestamp) > ttl_seconds:
        return None
    opportunities = [Opportunity.from_row(row) for row in entry.get("rows") or [] if isinstance(row, dict)]
    return opportunities, bool(entry.get("truncated"))


def _write_disk_search_cache(
    path: str | Path, cache_key: tuple[Any, ...], opportunities: list[Any], truncated: bool
) -> None:
    """Persist opportunities to a shared on-disk cache so a worker can warm the web app."""
    cache_path = Path(path)
    data: dict[str, Any] = {}
    if cache_path.exists():
        try:
            existing = json.loads(cache_path.read_text(encoding="utf-8"))
            if isinstance(existing, dict):
                data = existing
        except (json.JSONDecodeError, OSError):
            data = {}
    data[_disk_cache_key(cache_key)] = {
        "ts": time.time(),
        "truncated": bool(truncated),
        "rows": [opportunity.to_row() for opportunity in opportunities],
    }
    _atomic_write_text(cache_path, json.dumps(data, ensure_ascii=False, indent=2))


def _deep_search_opportunities(
    client: Any,
    keywords: list[str],
    max_pages: int = 20,
    paginate_by: int = 50,
    deadline_seconds: float | None = None,
) -> tuple[list[Any], bool]:
    """Fetch opportunities across pages, stopping early if the wall-clock deadline is hit.

    Returns ``(opportunities, truncated)`` where ``truncated`` is True when the deadline
    cut the crawl short so callers can surface partial-result advice.
    """
    by_ocid: dict[str, Any] = {}
    start = time.monotonic()
    truncated = False
    for keyword in keywords:
        if truncated:
            break
        for page in range(1, max_pages + 1):
            if deadline_seconds is not None and (time.monotonic() - start) > deadline_seconds:
                truncated = True
                break
            page_results = client.search_opportunities(keyword, page=page, paginate_by=paginate_by)
            if not page_results:
                break
            for opportunity in page_results:
                key = opportunity.ocid or f"{opportunity.process_code}|{opportunity.entity_id}|{opportunity.date}"
                by_ocid.setdefault(key, opportunity)
    return list(by_ocid.values()), truncated


def _apply_custom_variable_scoring(row: dict[str, Any], settings: dict[str, Any]) -> dict[str, Any]:
    text = " ".join(
        str(row.get(field) or "")
        for field in ("process_code", "entity_name", "description", "category", "procurement_method")
    ).lower()
    reasons = list(row.get("commercial_reasons") or [])
    score = int(row.get("commercial_score") or 0)
    for variable in settings.get("custom_variables") or []:
        name = str(variable.get("name") or "").strip()
        value = str(variable.get("value") or "").strip()
        if value and value.lower() in text:
            score = min(100, score + 5)
            reason = f"Coincide con variable {name}: {value}" if name else f"Coincide con variable personalizada: {value}"
            if reason not in reasons:
                reasons.append(reason)
    row["commercial_score"] = score
    row["commercial_reasons"] = reasons
    row["priority_label"] = "Alta" if score >= 70 else "Media" if score >= 35 else row.get("priority_label", "Baja")
    return row


def _text_matches(value: str | None, needle: str | None) -> bool:
    if not needle:
        return True
    return str(needle).strip().lower() in str(value or "").lower()


def _date_part(value: str | None) -> str:
    return str(value or "")[:10]


def _date_in_range(value: str | None, start: str | None, end: str | None) -> bool:
    current = _date_part(value)
    if not current:
        return not (start or end)
    if start and current < start:
        return False
    if end and current > end:
        return False
    return True


def _contract_object_matches(opportunity: Any, contract_object: str | None) -> bool:
    wanted = str(contract_object or "").strip().lower()
    if not wanted:
        return True
    category = str(getattr(opportunity, "category", "") or "").lower()
    text = " ".join(
        str(getattr(opportunity, field, "") or "")
        for field in ("process_code", "description", "procurement_method", "category")
    ).lower()
    if wanted == "bien":
        return category == "goods" or "bien" in text or "suministro" in text or "compra" in text
    if wanted == "servicio":
        return category == "services" and "consult" not in text and "supervisi" not in text and "expediente técnico" not in text
    if wanted == "obra":
        return category == "works" or "obra" in text or "construcción" in text or "mejoramiento" in text
    if wanted == "consultoria_obra":
        return "consult" in text or "supervisi" in text or "expediente técnico" in text
    return True


def _filter_opportunities(
    opportunities: list[Any],
    *,
    contract_object: str = "",
    entity_name: str = "",
    selection_type: str = "",
    description_filter: str = "",
    publication_from: str = "",
    publication_to: str = "",
    convocatoria_from: str = "",
    convocatoria_to: str = "",
) -> tuple[list[Any], int]:
    filtered: list[Any] = []
    for opportunity in opportunities:
        if not _contract_object_matches(opportunity, contract_object):
            continue
        if not _text_matches(getattr(opportunity, "entity_name", ""), entity_name):
            continue
        if not _text_matches(getattr(opportunity, "procurement_method", ""), selection_type):
            continue
        combined_description = " ".join([
            str(getattr(opportunity, "process_code", "") or ""),
            str(getattr(opportunity, "description", "") or ""),
        ])
        if not _text_matches(combined_description, description_filter):
            continue
        if not _date_in_range(getattr(opportunity, "date", ""), publication_from or None, publication_to or None):
            continue
        if not _date_in_range(getattr(opportunity, "tender_start_date", ""), convocatoria_from or None, convocatoria_to or None):
            continue
        filtered.append(opportunity)
    return filtered, len(opportunities) - len(filtered)


def _search_zero_result_advice(
    *,
    count: int,
    filtered_out_count: int,
    contract_object: str = "",
    entity_name: str = "",
    selection_type: str = "",
    description_filter: str = "",
    publication_from: str = "",
    publication_to: str = "",
    convocatoria_from: str = "",
    convocatoria_to: str = "",
    min_amount: float = 0,
) -> tuple[list[str], dict[str, Any]]:
    if count > 0:
        return [], {}
    advice: list[str] = []
    recommended = {
        "contract_object": contract_object,
        "entity_name": entity_name,
        "selection_type": selection_type,
        "description_filter": description_filter,
        "publication_from": publication_from,
        "publication_to": publication_to,
        "convocatoria_from": convocatoria_from,
        "convocatoria_to": convocatoria_to,
        "min_amount": min_amount,
    }
    if selection_type:
        advice.append("Tipo de selección puede venir vacío o con otra etiqueta en OECE/OCDS; déjalo vacío para descubrir primero.")
        recommended["selection_type"] = ""
    if convocatoria_from or convocatoria_to:
        advice.append("La fecha aproximada de convocatoria suele estar incompleta; limpia ese rango y filtra luego por publicación.")
        recommended["convocatoria_from"] = ""
        recommended["convocatoria_to"] = ""
    if description_filter and description_filter.strip().lower().endswith("s"):
        advice.append("Prueba la descripción en singular; por ejemplo, 'puente' en vez de 'puentes'.")
        recommended["description_filter"] = description_filter.strip()[:-1]
    if publication_from or publication_to:
        advice.append("Si sigues en cero, elimina temporalmente el rango de publicación y revisa resultados históricos.")
    if min_amount:
        advice.append("Baja el monto mínimo a 0 para descubrir oportunidades y súbelo después.")
        recommended["min_amount"] = 0
    if filtered_out_count:
        advice.append(f"Los filtros ocultaron {filtered_out_count} resultado(s); amplia filtros antes de concluir que no hay oportunidades.")
    if not advice:
        advice.append("No hubo coincidencias. Prueba una palabra más general como puente, obra vial o carretera.")
    return advice, recommended


def create_app(
    dashboard_path: str | Path = DEFAULT_DASHBOARD_PATH,
    static_dir: str | Path = DEFAULT_STATIC_DIR,
    tracking_db_path: str | Path = DEFAULT_TRACKING_DB_PATH,
    settings_path: str | Path = DEFAULT_SETTINGS_PATH,
    seace_client: Any | None = None,
    ficha_capture_service: Any | None = None,
    api_key: str | None = None,
    search_cache_ttl: float = DEFAULT_SEARCH_CACHE_TTL,
    search_deadline: float = DEFAULT_SEARCH_DEADLINE,
    search_cache_path: str | Path | None = None,
    supabase_url: str | None = None,
    supabase_anon_key: str | None = None,
    supabase_verify: Any | None = None,
) -> FastAPI:
    app = FastAPI(title="LicitaScan", version="0.1.0")
    dashboard_path = Path(dashboard_path)
    static_dir = Path(static_dir)
    tracking_db_path = Path(tracking_db_path)
    settings_path = Path(settings_path)
    api_client = seace_client or SeaceApiClient()
    capture_service = ficha_capture_service or _default_ficha_capture_service
    configured_api_key = (api_key if api_key is not None else os.getenv("LICITASCAN_API_KEY", "")) or ""
    supabase_auth = SupabaseAuth(supabase_url, supabase_anon_key, verify=supabase_verify)
    search_cache = _TTLCache(search_cache_ttl)
    disk_search_cache_path = Path(search_cache_path) if search_cache_path is not None else None

    # Endpoints abiertos (sin auth): health para el monitor y config para que el
    # frontend pueda inicializar Supabase antes de tener sesión.
    PUBLIC_API_PATHS = {"/api/health", "/api/config"}

    @app.middleware("http")
    async def require_auth(request: Request, call_next):  # type: ignore[no-untyped-def]
        # Gate de /api/*: acepta un Bearer token de Supabase (login con Google) o,
        # como compatibilidad/admin, la cabecera X-API-Key. Si no hay NADA
        # configurado (ni Supabase ni API key), la auth queda desactivada (dev/local).
        path = request.url.path
        if path.startswith("/api/") and path not in PUBLIC_API_PATHS:
            user: dict[str, Any] | None = None
            if supabase_auth.enabled:
                token = bearer_token(request)
                if token:
                    user = supabase_auth.verify_token(token)
            if user is None and configured_api_key:
                provided = request.headers.get("X-API-Key", "")
                if hmac.compare_digest(provided, configured_api_key):
                    user = {"id": "api-key", "email": ""}
            if user is None and (supabase_auth.enabled or configured_api_key):
                return JSONResponse(
                    {"detail": "No autenticado"},
                    status_code=401,
                    headers={"WWW-Authenticate": "Bearer"},
                )
            request.state.user = user or {}
        return await call_next(request)

    @app.middleware("http")
    async def add_security_headers(request: Request, call_next):  # type: ignore[no-untyped-def]
        response = await call_next(request)
        for key, value in SECURITY_HEADERS.items():
            response.headers[key] = value
        return response

    @app.get("/api/health")
    def health() -> dict[str, str]:
        return {"status": "ok"}

    @app.get("/api/config")
    def public_config() -> dict[str, Any]:
        # Datos públicos para que el frontend inicialice Supabase (la anon key es
        # pública por diseño). Si auth_enabled es False, el front no exige login.
        return {
            "auth_enabled": supabase_auth.enabled,
            "supabase_url": supabase_auth.url,
            "supabase_anon_key": supabase_auth.anon_key,
        }

    @app.get("/api/me")
    def current_user(request: Request) -> dict[str, Any]:
        return {"user": getattr(request.state, "user", {}) or {}}

    @app.get("/api/dashboard")
    def dashboard() -> dict[str, Any]:
        return load_dashboard(dashboard_path)

    @app.get("/api/settings")
    def get_settings() -> dict[str, Any]:
        return load_settings(settings_path)

    @app.put("/api/settings")
    async def put_settings(request: Request) -> dict[str, Any]:
        payload = await request.json()
        if not isinstance(payload, dict):
            raise HTTPException(status_code=400, detail="settings payload must be an object")
        return save_settings(payload, settings_path)

    @app.post("/api/dismiss")
    async def dismiss_opportunity(request: Request) -> dict[str, Any]:
        payload = await request.json()
        if not isinstance(payload, dict):
            raise HTTPException(status_code=400, detail="dismiss payload must be an object")
        ocid = str(payload.get("ocid") or "").strip()
        if not ocid:
            raise HTTPException(status_code=400, detail="ocid is required")
        def _add_ignored(settings: dict[str, Any]) -> dict[str, Any]:
            settings["ignored_ocids"] = _unique_text_list([*(settings.get("ignored_ocids") or []), ocid])
            return settings
        return update_settings(_add_ignored, settings_path)

    @app.post("/api/dismiss/restore")
    async def restore_dismissed_opportunity(request: Request) -> dict[str, Any]:
        payload = await request.json()
        if not isinstance(payload, dict):
            raise HTTPException(status_code=400, detail="restore payload must be an object")
        ocid = str(payload.get("ocid") or "").strip()
        if not ocid:
            raise HTTPException(status_code=400, detail="ocid is required")
        def _remove_ignored(settings: dict[str, Any]) -> dict[str, Any]:
            settings["ignored_ocids"] = [item for item in settings.get("ignored_ocids", []) if item != ocid]
            return settings
        return update_settings(_remove_ignored, settings_path)

    @app.get("/api/search")
    def search_opportunities(
        keywords: str = "",
        min_amount: float | None = None,
        max_pages: int = 20,
        paginate_by: int = 50,
        result_limit: int = 100,
        contract_object: str = "",
        entity_name: str = "",
        selection_type: str = "",
        description_filter: str = "",
        publication_from: str = "",
        publication_to: str = "",
        convocatoria_from: str = "",
        convocatoria_to: str = "",
    ) -> dict[str, Any]:
        settings = load_settings(settings_path)
        clean_keywords = _parse_keywords(keywords) if keywords.strip() else list(settings["keywords"])
        if not clean_keywords:
            raise HTTPException(status_code=400, detail="keywords is required")
        effective_min_amount = settings["min_amount"] if min_amount is None else min_amount
        safe_max_pages = min(max(1, max_pages), 50)
        safe_paginate_by = min(max(1, paginate_by), 100)
        cache_key = (tuple(sorted(clean_keywords)), safe_max_pages, safe_paginate_by)
        cached = search_cache.get(cache_key)
        if cached is None and disk_search_cache_path is not None:
            cached = _read_disk_search_cache(disk_search_cache_path, cache_key, search_cache_ttl)
        if cached is not None:
            opportunities, search_truncated = cached
            search_cache.set(cache_key, cached)
            from_cache = True
        else:
            opportunities, search_truncated = _deep_search_opportunities(
                api_client,
                clean_keywords,
                max_pages=safe_max_pages,
                paginate_by=safe_paginate_by,
                deadline_seconds=search_deadline,
            )
            search_cache.set(cache_key, (opportunities, search_truncated))
            if disk_search_cache_path is not None:
                _write_disk_search_cache(disk_search_cache_path, cache_key, opportunities, search_truncated)
            from_cache = False
        ignored_ocids = set(settings.get("ignored_ocids") or [])
        visible_opportunities = [opportunity for opportunity in opportunities if opportunity.ocid not in ignored_ocids]
        ignored_count = len(opportunities) - len(visible_opportunities)
        filtered_by_seace, filtered_out_count = _filter_opportunities(
            visible_opportunities,
            contract_object=contract_object,
            entity_name=entity_name,
            selection_type=selection_type,
            description_filter=description_filter,
            publication_from=publication_from,
            publication_to=publication_to,
            convocatoria_from=convocatoria_from,
            convocatoria_to=convocatoria_to,
        )
        filtered = [opportunity for opportunity in filtered_by_seace if opportunity.amount is None or opportunity.amount >= effective_min_amount]
        negative_keywords = settings.get("negative_keywords") or []
        rows = [
            score_relevance(
                _apply_custom_variable_scoring(_search_result_row(opportunity), settings),
                clean_keywords,
                negative_keywords,
            )
            for opportunity in filtered
        ]
        # Etapa 1: ordena primero por cercanía a las keywords (relevancia), luego por
        # señal comercial y monto como desempate. No se descarta nada: se muestra todo.
        rows.sort(
            key=lambda item: (
                item.get("relevance_score") or 0,
                item.get("commercial_score") or 0,
                item.get("amount") or 0,
            ),
            reverse=True,
        )
        total_found = len(rows)
        safe_result_limit = min(max(1, result_limit), 500)
        visible_rows = rows[:safe_result_limit]
        search_advice, recommended_relaxation = _search_zero_result_advice(
            count=len(visible_rows),
            filtered_out_count=filtered_out_count,
            contract_object=contract_object,
            entity_name=entity_name,
            selection_type=selection_type,
            description_filter=description_filter,
            publication_from=publication_from,
            publication_to=publication_to,
            convocatoria_from=convocatoria_from,
            convocatoria_to=convocatoria_to,
            min_amount=effective_min_amount,
        )
        return {
            "count": len(visible_rows),
            "total_found": total_found,
            "results": visible_rows,
            "keywords": clean_keywords,
            "min_amount": effective_min_amount,
            "searched_pages_limit": safe_max_pages,
            "from_cache": from_cache,
            "search_truncated": search_truncated,
            "ignored_count": ignored_count,
            "filtered_out_count": filtered_out_count,
            "search_advice": search_advice,
            "recommended_relaxation": recommended_relaxation,
            "filters": {
                "contract_object": contract_object,
                "entity_name": entity_name,
                "selection_type": selection_type,
                "description_filter": description_filter,
                "publication_from": publication_from,
                "publication_to": publication_to,
                "convocatoria_from": convocatoria_from,
                "convocatoria_to": convocatoria_to,
            },
        }

    @app.post("/api/track")
    async def track_opportunities(request: Request) -> dict[str, Any]:
        payload = await request.json()
        ocids = [str(item).strip() for item in payload.get("ocids", []) if str(item).strip()]
        if not ocids:
            raise HTTPException(status_code=400, detail="ocids is required")
        store = TrackingStore(tracking_db_path)
        store.initialize()
        events = sync_ocids(api_client, store, ocids, dashboard_path=dashboard_path)
        return {"tracked": ocids, "events": len(events), "dashboard": str(dashboard_path)}

    @app.get("/api/opportunities/{ocid}")
    def opportunity_detail(ocid: str) -> dict[str, Any]:
        data = load_dashboard(dashboard_path)
        opportunity = _find_opportunity(data, ocid)
        related_events = [event for event in data["recent_events"] if event.get("ocid") == ocid]
        return {
            "opportunity": deepcopy(opportunity),
            "events": related_events,
            "timeline": _build_timeline(opportunity, related_events),
        }

    @app.get("/api/opportunities/{ocid}/ficha")
    def opportunity_ficha(ocid: str) -> dict[str, Any]:
        data = load_dashboard(dashboard_path)
        opportunity = _find_opportunity(data, ocid)
        return _ficha_payload(opportunity)

    @app.post("/api/opportunities/{ocid}/ficha/capture")
    def opportunity_ficha_capture(ocid: str) -> dict[str, Any]:
        data = load_dashboard(dashboard_path)
        opportunity = _find_opportunity(data, ocid)
        evidence_dir = static_dir / "evidencias"
        try:
            return capture_service(opportunity, evidence_dir)
        except Exception as exc:  # noqa: BLE001 - user-facing automation failure
            raise HTTPException(status_code=502, detail=f"No se pudo automatizar la ficha SEACE: {exc}") from exc

    @app.get("/api/opportunities/{ocid}/documents")
    def opportunity_documents(ocid: str, verify: bool = True, analyze: bool = False) -> dict[str, Any]:
        data = load_dashboard(dashboard_path)
        opportunity = _find_opportunity(data, ocid)
        documents = _documents_from_record_or_dashboard(api_client, opportunity)
        enriched = [_document_with_status(document, verify=verify, analyze=analyze) for document in documents]
        return {
            "ocid": ocid,
            "count": len(enriched),
            "documents": enriched,
            "quota_policy": {
                "viewing_consumes_credit": False,
                "verification_consumes_download": False,
                "download_consumes_only_on_success": True,
                "failed_download_consumes_download": False,
            },
            "browser_help": "Algunos documentos SEACE pueden requerir configuración de navegador; LicitaScan verifica el enlace antes de descontar descargas.",
        }

    @app.get("/api/opportunities/{ocid}/eto")
    def opportunity_technical_file(ocid: str, verify: bool = True, analyze: bool = False) -> dict[str, Any]:
        data = load_dashboard(dashboard_path)
        opportunity = _find_opportunity(data, ocid)
        record_payload = api_client.get_record(ocid) if api_client else {}
        if not record_payload:
            record_payload = {
                "compiledRelease": {
                    "tender": {
                        "title": opportunity.get("process_code", ""),
                        "description": opportunity.get("description", ""),
                        "mainProcurementCategory": opportunity.get("category", ""),
                        "documents": opportunity.get("official_documents", []),
                    }
                }
            }
        return build_technical_file_response(
            record_payload,
            verify=verify,
            analyze=analyze,
            verifier=verify_document_link,
            analyzer=analyze_document,
        )

    @app.get("/api/documents/download")
    def download_document(url: str, filename: str = "") -> Response:
        known_document = _find_document_for_url(api_client, dashboard_path, url)
        suggested_name = filename or str((known_document or {}).get("suggested_filename") or (known_document or {}).get("title") or "documento-oficial.bin")
        try:
            document = download_verified_document(url, suggested_name)
        except Exception as exc:  # noqa: BLE001 - user-facing no-quota failure
            raise HTTPException(
                status_code=502,
                detail=f"No se pudo descargar el documento oficial; no debe consumir cuota de descarga. {exc}",
            ) from exc
        return Response(
            content=document.content,
            media_type=document.content_type,
            headers={"Content-Disposition": f'attachment; filename="{document.filename}"'},
        )

    @app.get("/api/opportunities/{ocid}/export")
    def export_opportunity(ocid: str) -> JSONResponse:
        data = load_dashboard(dashboard_path)
        opportunity = _find_opportunity(data, ocid)
        related_events = [event for event in data["recent_events"] if event.get("ocid") == ocid]
        filename = f"expediente-{_safe_filename(str(opportunity.get('process_code') or ocid))}.json"
        return JSONResponse(
            content=_expediente_export(opportunity, related_events),
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    @app.get("/api/market-intel")
    def market_intel(keyword: str = "", year: int | None = None, force_refresh: bool = False) -> dict[str, Any]:
        """Return CONOSCE market intelligence: top entities, categories and winners."""
        return fetch_market_intel(keyword=keyword, year=year, force_refresh=force_refresh)

    @app.get("/api/export/xlsx")
    def export_xlsx() -> Response:
        """Export the current tracking dashboard as an Excel workbook."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise HTTPException(status_code=501, detail="openpyxl no está instalado; instala con pip install openpyxl")

        data = load_dashboard(dashboard_path)
        opportunities = data.get("opportunities", [])

        wb = openpyxl.Workbook()

        # ── Sheet 1: Oportunidades ───────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Oportunidades"
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(fill_type="solid", fgColor="0B43A5")
        headers = [
            "OCID", "Código proceso", "Entidad", "Descripción",
            "Categoría", "Monto ref. (S/)", "Moneda",
            "Etapa", "Resultado", "Fecha crítica",
            "Ganador", "RUC ganador", "Monto adj. (S/)", "Fecha adjudicación",
            "Contrato ID", "Fecha contrato",
            "Score comercial", "Prioridad", "Acción recomendada",
        ]
        ws1.append(headers)
        for cell in ws1[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for opp in opportunities:
            ws1.append([
                opp.get("ocid", ""),
                opp.get("process_code", ""),
                opp.get("entity_name", ""),
                opp.get("description", ""),
                opp.get("category", ""),
                opp.get("amount") or "",
                opp.get("currency", ""),
                opp.get("stage", ""),
                opp.get("outcome", ""),
                (opp.get("next_critical_date") or "")[:10],
                opp.get("winner_name", ""),
                opp.get("winner_ruc", ""),
                opp.get("awarded_amount") or "",
                (opp.get("award_date") or "")[:10],
                opp.get("contract_id", ""),
                (opp.get("contract_date_signed") or "")[:10],
                opp.get("commercial_score") or "",
                opp.get("priority_label", ""),
                opp.get("recommended_action", ""),
            ])

        # Auto-fit columns (approximate)
        for col in ws1.columns:
            max_length = max((len(str(cell.value or "")) for cell in col), default=0)
            ws1.column_dimensions[col[0].column_letter].width = min(max_length + 4, 50)

        # ── Sheet 2: Eventos recientes ───────────────────────────────────────
        ws2 = wb.create_sheet("Eventos recientes")
        event_headers = ["OCID", "Tipo evento", "Título", "Mensaje", "Severidad", "Fecha"]
        ws2.append(event_headers)
        for cell in ws2[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for event in data.get("recent_events", []):
            ws2.append([
                event.get("ocid", ""),
                event.get("event_type", ""),
                event.get("title", ""),
                event.get("message", ""),
                event.get("severity", ""),
                (event.get("occurred_at") or "")[:19],
            ])

        # ── Sheet 3: Resumen ─────────────────────────────────────────────────
        ws3 = wb.create_sheet("Resumen")
        ws3.append(["Etapa", "Cantidad"])
        for stage, count in data.get("counts_by_stage", {}).items():
            ws3.append([stage, count])
        ws3.append([])
        ws3.append(["Resultado", "Cantidad"])
        for outcome, count in data.get("counts_by_outcome", {}).items():
            ws3.append([outcome, count])

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return Response(
            content=buffer.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=\"licitascan-export.xlsx\""},
        )

    if static_dir.exists():
        app.mount("/assets", StaticFiles(directory=static_dir), name="assets")

        @app.get("/")
        def index() -> FileResponse:
            return FileResponse(static_dir / "index.html")

    return app


app = create_app(search_cache_path=DEFAULT_SEARCH_CACHE_PATH)
